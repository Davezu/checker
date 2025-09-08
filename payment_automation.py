import json
import tkinter as tk
from tkinter import messagebox, simpledialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time


class PaymentAutomation:
    def __init__(self):
        self.config = self.load_config()
        self.driver = None
        self.selected_card = None
        
    def load_config(self):
        """Load configuration from config.json file"""
        try:
            with open('config.json', 'r') as file:
                return json.load(file)
        except FileNotFoundError:
            messagebox.showerror("Error", "config.json file not found!")
            return None
        except json.JSONDecodeError:
            messagebox.showerror("Error", "Invalid JSON format in config.json!")
            return None
    
    def show_card_selection(self):
        """Display a popup window for card selection"""
        if not self.config:
            return None
            
        root = tk.Tk()
        root.title("Select Payment Card")
        root.geometry("400x300")
        root.resizable(False, False)
        
        # Center the window
        root.update_idletasks()
        x = (root.winfo_screenwidth() // 2) - (400 // 2)
        y = (root.winfo_screenheight() // 2) - (300 // 2)
        root.geometry(f"400x300+{x}+{y}")
        
        selected_card = None
        
        def on_card_select(card):
            nonlocal selected_card
            selected_card = card
            root.destroy()
        
        def on_cancel():
            root.destroy()
        
        # Title
        title_label = tk.Label(root, text="Choose a Payment Card", 
                              font=("Arial", 16, "bold"), pady=20)
        title_label.pack()
        
        # Card buttons
        button_frame = tk.Frame(root)
        button_frame.pack(pady=20)
        
        for i, card in enumerate(self.config['cards']):
            card_text = f"{card['name']}\n****-****-****-{card['number'][-4:]}"
            btn = tk.Button(button_frame, text=card_text, width=25, height=3,
                           command=lambda c=card: on_card_select(c),
                           font=("Arial", 10), relief="raised", bd=2)
            btn.pack(pady=5)
        
        # Cancel button
        cancel_btn = tk.Button(root, text="Cancel", command=on_cancel,
                              bg="#ff4444", fg="white", font=("Arial", 10))
        cancel_btn.pack(pady=20)
        
        root.mainloop()
        return selected_card
    
    def get_payment_amount(self):
        """Get payment amount from user"""
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        
        amount = simpledialog.askfloat("Payment Amount", 
                                      "Enter the payment amount:",
                                      minvalue=0.01, maxvalue=99999.99)
        root.destroy()
        return amount
    
    def setup_driver(self):
        """Setup WebDriver with specified options - tries Firefox first, then Chrome"""
        
        # Try Firefox first (more reliable)
        try:
            print("Setting up Firefox WebDriver...")
            firefox_options = FirefoxOptions()
            
            if self.config['browser_settings']['headless']:
                firefox_options.add_argument('--headless')
            
            firefox_options.add_argument('--disable-blink-features=AutomationControlled')
            firefox_options.set_preference("dom.webdriver.enabled", False)
            firefox_options.set_preference('useAutomationExtension', False)
            
            window_size = self.config['browser_settings']['window_size']
            firefox_options.add_argument(f'--width={window_size[0]}')
            firefox_options.add_argument(f'--height={window_size[1]}')
            
            service = FirefoxService(GeckoDriverManager().install())
            self.driver = webdriver.Firefox(service=service, options=firefox_options)
            print("Firefox WebDriver initialized successfully!")
            
        except Exception as e:
            print(f"Firefox setup failed: {e}")
            print("Trying Chrome as fallback...")
            
            # Fallback to Chrome
            chrome_options = Options()
            
            if self.config['browser_settings']['headless']:
                chrome_options.add_argument('--headless')
            
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-web-security')
            chrome_options.add_argument('--allow-running-insecure-content')
            
            window_size = self.config['browser_settings']['window_size']
            chrome_options.add_argument(f'--window-size={window_size[0]},{window_size[1]}')
            
            try:
                # First, try to use local ChromeDriver if it exists
                local_chromedriver = os.path.join(os.getcwd(), "chromedriver.exe")
                if os.path.exists(local_chromedriver):
                    print("Using local ChromeDriver...")
                    service = Service(local_chromedriver)
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                else:
                    # Try to get ChromeDriver automatically
                    print("Downloading ChromeDriver automatically...")
                    service = Service(ChromeDriverManager().install())
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
            except Exception as e2:
                print(f"Chrome setup also failed: {e2}")
                print("\nSolutions:")
                print("1. Install Firefox: https://www.mozilla.org/firefox/")
                print("2. Or downgrade Chrome to version 131 or lower")
                print("3. Or run: python setup_chromedriver.py")
                raise Exception("Could not initialize any WebDriver.")
        
        # Set anti-detection properties
        try:
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        except:
            pass  # Firefox might not support this
            
        implicit_wait = self.config['browser_settings']['implicit_wait']
        self.driver.implicitly_wait(implicit_wait)
    
    def login_to_foodpanda(self):
        """Login to FoodPanda if required"""
        try:
            print("Checking if login is required...")
            
            # Check if we're on login page or need to login
            if "login" in self.driver.current_url.lower() or self.driver.find_elements(By.CSS_SELECTOR, self.config['website']['selectors']['login_email']):
                print("Login required. Logging in...")
                
                # Fill email
                email_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, self.config['website']['selectors']['login_email']))
                )
                email_field.clear()
                email_field.send_keys(self.config['foodpanda_login']['email'])
                
                # Fill password
                password_field = self.driver.find_element(
                    By.CSS_SELECTOR, self.config['website']['selectors']['login_password'])
                password_field.clear()
                password_field.send_keys(self.config['foodpanda_login']['password'])
                
                # Click login button
                login_btn = self.driver.find_element(
                    By.CSS_SELECTOR, self.config['website']['selectors']['login_button'])
                login_btn.click()
                
                # Wait for login to complete
                time.sleep(3)
                print("Login completed!")
                
            return True
            
        except Exception as e:
            print(f"Login process failed: {str(e)}")
            return False

    def fill_payment_form(self, amount):
        """Fill out the FoodPanda Philippines PandaPay top-up form"""
        try:
            # Navigate to FoodPanda PandaPay top-up page
            print(f"Navigating to {self.config['website']['url']}...")
            self.driver.get(self.config['website']['url'])
            
            # Wait for the page to load
            print("Waiting for FoodPanda page to load...")
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 
                    self.config['website']['selectors']['amount_input']))
            )
            
            print("✅ FoodPanda PandaPay page loaded successfully!")
            
            # Fill top-up amount
            print(f"Entering top-up amount: ₱{amount}")
            amount_field = self.driver.find_element(
                By.CSS_SELECTOR, self.config['website']['selectors']['amount_input'])
            amount_field.clear()
            amount_field.send_keys(str(amount))
            
            # Wait for amount to be processed
            time.sleep(2)
            
            # Select Credit/Debit Card payment method if not already selected
            try:
                print("Selecting Credit/Debit Card payment method...")
                credit_card_radio = self.driver.find_element(
                    By.CSS_SELECTOR, self.config['website']['selectors']['credit_card_option'])
                if not credit_card_radio.is_selected():
                    credit_card_radio.click()
                    time.sleep(1)
            except:
                print("Credit card option already selected or not found...")
            
            # Wait for card form to appear and fill card details
            print("Filling card details...")
            
            # Fill card number
            card_number_field = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 
                    self.config['website']['selectors']['card_number']))
            )
            card_number_field.clear()
            card_number_field.send_keys(self.selected_card['number'])
            time.sleep(1)
            
            # Fill expiry date (MM/YY format)
            expiry_field = self.driver.find_element(
                By.CSS_SELECTOR, self.config['website']['selectors']['expiry_date'])
            expiry_field.clear()
            # Convert MM/YY format (e.g., "12/25" stays as "12/25")
            expiry_formatted = self.selected_card['expiry']
            expiry_field.send_keys(expiry_formatted)
            time.sleep(1)
            
            # Fill CVC
            cvc_field = self.driver.find_element(
                By.CSS_SELECTOR, self.config['website']['selectors']['cvc'])
            cvc_field.clear()
            cvc_field.send_keys(self.selected_card['cvc'])
            time.sleep(1)
            
            # Fill cardholder name
            print("Filling cardholder name...")
            cardholder_field = self.driver.find_element(
                By.CSS_SELECTOR, self.config['website']['selectors']['cardholder_name'])
            cardholder_field.clear()
            cardholder_field.send_keys(self.selected_card.get('holder_name', 'John Doe'))
            
            # Wait for all fields to be processed
            time.sleep(3)
            
            # Look for and click the Pay button
            print("Looking for Pay button...")
            try:
                # Try to find Pay button by text
                pay_button = self.driver.find_element(By.XPATH, "//button[contains(text(), 'Pay')]")
                pay_button.click()
                print("Clicked Pay button!")
            except:
                # Fallback to submit button
                try:
                    submit_btn = self.driver.find_element(
                        By.CSS_SELECTOR, self.config['website']['selectors']['submit_button'])
                    submit_btn.click()
                    print("Clicked Submit button!")
                except:
                    print("Could not find Pay or Submit button")
                    return False
            
            return True
            
        except Exception as e:
            print(f"Error filling FoodPanda form: {str(e)}")
            return False
    
    def capture_confirmation(self):
        """Capture confirmation message and transaction ID"""
        try:
            # Wait for confirmation page to load
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 
                    self.config['website']['selectors']['confirmation_message']))
            )
            
            # Get confirmation message
            confirmation_element = self.driver.find_element(
                By.CSS_SELECTOR, self.config['website']['selectors']['confirmation_message'])
            confirmation_message = confirmation_element.text
            
            # Get transaction ID
            transaction_id = "N/A"
            try:
                transaction_element = self.driver.find_element(
                    By.CSS_SELECTOR, self.config['website']['selectors']['transaction_id'])
                transaction_id = transaction_element.text
            except:
                print("Transaction ID element not found, using 'N/A'")
            
            return confirmation_message, transaction_id
            
        except Exception as e:
            print(f"Error capturing confirmation: {str(e)}")
            return "Error capturing confirmation", "N/A"
    
    def save_to_excel(self, amount, confirmation_message, transaction_id):
        """Save transaction details to Excel file"""
        excel_file = self.config['excel_file']
        
        # Create or load workbook
        if os.path.exists(excel_file):
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            # Add headers
            headers = ['Date', 'Time', 'Card Used', 'Amount', 'Transaction ID', 'Confirmation Message']
            worksheet.append(headers)
        
        # Prepare data
        current_time = datetime.now()
        date_str = current_time.strftime("%Y-%m-%d")
        time_str = current_time.strftime("%H:%M:%S")
        card_name = self.selected_card['name']
        
        # Add transaction data
        row_data = [date_str, time_str, card_name, amount, transaction_id, confirmation_message]
        worksheet.append(row_data)
        
        # Save workbook
        workbook.save(excel_file)
        print(f"Transaction saved to {excel_file}")
    
    def run(self):
        """Main execution method"""
        if not self.config:
            return
        
        try:
            # Show card selection popup
            self.selected_card = self.show_card_selection()
            if not self.selected_card:
                print("No card selected. Exiting...")
                return
            
            # Get payment amount
            amount = self.get_payment_amount()
            if not amount:
                print("No amount entered. Exiting...")
                return
            
            print(f"Selected card: {self.selected_card['name']}")
            print(f"Payment amount: ${amount}")
            
            # Setup WebDriver
            print("Setting up browser...")
            self.setup_driver()
            
            # Fill payment form
            print("Filling payment form...")
            if self.fill_payment_form(amount):
                print("Form submitted successfully!")
                
                # Capture confirmation
                print("Capturing confirmation...")
                confirmation_message, transaction_id = self.capture_confirmation()
                
                # Save to Excel
                print("Saving transaction to Excel...")
                self.save_to_excel(amount, confirmation_message, transaction_id)
                
                print("Payment automation completed successfully!")
                messagebox.showinfo("Success", 
                    f"Payment processed successfully!\nTransaction ID: {transaction_id}")
            else:
                print("Failed to fill payment form.")
                messagebox.showerror("Error", "Failed to fill payment form.")
                
        except Exception as e:
            print(f"An error occurred: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            
        finally:
            # Close browser
            if self.driver:
                print("Closing browser...")
                self.driver.quit()


def main():
    """Main entry point"""
    print("Starting Payment Automation...")
    app = PaymentAutomation()
    app.run()


if __name__ == "__main__":
    main()
